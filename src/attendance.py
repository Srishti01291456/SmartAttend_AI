import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


ATTENDANCE_FOLDER = "attendance"

ATTENDANCE_FILE = os.path.join(
    ATTENDANCE_FOLDER,
    "attendance.xlsx"
)


def create_excel_file():

    os.makedirs(
        ATTENDANCE_FOLDER,
        exist_ok=True
    )


    if not os.path.exists(ATTENDANCE_FILE):

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Attendance"


        sheet.append(
            [
                "Student ID",
                "Name",
                "Date",
                "Time"
            ]
        )


        # Header formatting

        for cell in sheet[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )


        # Column width

        sheet.column_dimensions["A"].width = 15
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 15


        # Freeze header

        sheet.freeze_panes = "A2"


        workbook.save(
            ATTENDANCE_FILE
        )

        workbook.close()



def mark_attendance(student_id, name):


    create_excel_file()


    today = datetime.now().strftime(
        "%Y-%m-%d"
    )


    current_time = datetime.now().strftime(
        "%H:%M:%S"
    )


    workbook = load_workbook(
        ATTENDANCE_FILE
    )


    sheet = workbook["Attendance"]



    # Duplicate checking

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):


        if (
            str(row[0]) == str(student_id)
            and row[2] == today
        ):

            workbook.close()

            return "Already Marked Today"



    # Add new attendance

    sheet.append(
        [
            student_id,
            name,
            today,
            current_time
        ]
    )


    # Apply border to new row

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )


    for cell in sheet[sheet.max_row]:

        cell.border = thin_border

        cell.alignment = Alignment(
            horizontal="center"
        )


    workbook.save(
        ATTENDANCE_FILE
    )


    workbook.close()


    print(
        "Attendance marked:",
        name
    )


    return "Attendance Marked"